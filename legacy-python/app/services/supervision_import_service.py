from abc import ABC, abstractmethod
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import ResourceNotFoundError
from app.db import get_connection
from app.repositories.excel_import_template_repository import (
    get_enabled_template,
    list_enabled_templates,
)
from app.repositories.supervision_batch_repository import create_batch, finish_batch
from app.repositories.supervision_item_repository import create_item
from app.services.supervision_item_service import generate_item_no


DEFAULT_TEMPLATE_CODE = "standard_supervision"


class ExcelImportStrategy(ABC):
    handler_code: str

    @abstractmethod
    def read_rows(self, content: bytes, template: dict[str, Any]) -> list[dict[str, Any]]:
        raise NotImplementedError

    @abstractmethod
    def build_item_payload(
        self,
        *,
        row: dict[str, Any],
        row_number: int,
        batch_id: str,
        created_by: str,
    ) -> dict[str, Any]:
        raise NotImplementedError


class StandardSupervisionExcelStrategy(ExcelImportStrategy):
    handler_code = "standard_supervision"

    def read_rows(self, content: bytes, template: dict[str, Any]) -> list[dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active
        headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        field_indexes = self.resolve_field_indexes(headers, template)
        rows: list[dict[str, Any]] = []

        for row_number, excel_row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if all(cell.value in (None, "") for cell in excel_row):
                continue
            item: dict[str, Any] = {"source_row_no": row_number}
            for field, index in field_indexes.items():
                item[field] = excel_row[index].value if index < len(excel_row) else None
            rows.append(item)
        return rows

    def build_item_payload(
        self,
        *,
        row: dict[str, Any],
        row_number: int,
        batch_id: str,
        created_by: str,
    ) -> dict[str, Any]:
        title = clean_text(row.get("title"))
        if not title:
            raise ValueError("缺少事项标题")

        return {
            "batch_id": batch_id,
            "item_no": clean_text(row.get("item_no")) or generate_item_no(),
            "title": title,
            "description": clean_text(row.get("description")) or None,
            "source_row_no": row_number,
            "priority": clean_text(row.get("priority")) or "normal",
            "status": clean_text(row.get("status")) or "pending_assign",
            "deadline_at": parse_datetime(row.get("deadline_at")),
            "created_by": clean_text(row.get("created_by")) or created_by,
        }

    def resolve_field_indexes(
        self,
        headers: list[str],
        template: dict[str, Any],
    ) -> dict[str, int]:
        indexes: dict[str, int] = {}
        field_aliases = normalize_template_mapping(template)
        for index, header in enumerate(headers):
            for field, aliases in field_aliases.items():
                if header in set(aliases):
                    indexes[field] = index
        return indexes


STRATEGIES: dict[str, ExcelImportStrategy] = {
    StandardSupervisionExcelStrategy.handler_code: StandardSupervisionExcelStrategy(),
}


def list_import_templates() -> list[dict[str, Any]]:
    with get_connection() as conn:
        return list_enabled_templates(conn)


def import_supervision_excel(
    *,
    filename: str,
    content: bytes,
    created_by: str,
    template_code: str = DEFAULT_TEMPLATE_CODE,
) -> dict[str, Any]:
    with get_connection() as conn:
        template = get_enabled_template(conn, template_code)

    strategy = get_strategy(template["handler_code"])
    batch_no = f"BATCH-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    batch_name = f"{template['template_name']} - {filename}" if filename else batch_no
    errors: list[str] = []

    with get_connection() as conn:
        batch = create_batch(
            conn,
            batch_no=batch_no,
            batch_name=batch_name,
            source_file_id=filename,
            created_by=created_by,
        )

    rows = strategy.read_rows(content, template)
    total_count = len(rows)
    success_count = 0

    for row in rows:
        row_number = int(row.get("source_row_no") or 0)
        try:
            payload = strategy.build_item_payload(
                row=row,
                row_number=row_number,
                batch_id=batch["id"],
                created_by=created_by,
            )
            with get_connection() as conn:
                create_item(conn, payload)
            success_count += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"第 {row_number} 行导入失败：{exc}")

    failed_count = total_count - success_count
    import_status = "completed" if failed_count == 0 else "failed"
    with get_connection() as conn:
        batch = finish_batch(
            conn,
            batch["id"],
            import_status=import_status,
            total_count=total_count,
            success_count=success_count,
            failed_count=failed_count,
        )

    return {
        "batch_id": batch["id"],
        "batch_no": batch["batch_no"],
        "batch_name": batch["batch_name"],
        "import_status": batch["import_status"],
        "total_count": batch["total_count"],
        "success_count": batch["success_count"],
        "failed_count": batch["failed_count"],
        "errors": errors,
    }


def get_strategy(handler_code: str) -> ExcelImportStrategy:
    try:
        return STRATEGIES[handler_code]
    except KeyError as exc:
        raise ResourceNotFoundError(f"excel import handler not found: {handler_code}") from exc


def normalize_template_mapping(template: dict[str, Any]) -> dict[str, list[str]]:
    source_columns = template.get("source_columns") or []
    entity_fields = template.get("entity_fields") or []
    aliases: dict[str, list[str]] = {}

    for source_column, entity_field in zip(source_columns, entity_fields, strict=False):
        clean_entity_field = clean_text(entity_field)
        clean_source_column = clean_text(source_column)
        if clean_entity_field and clean_source_column:
            aliases[clean_entity_field] = [clean_source_column]

    mapping_config = template.get("mapping_config") or {}
    extra_aliases = mapping_config.get("aliases") or {}
    if isinstance(extra_aliases, dict):
        for entity_field, columns in extra_aliases.items():
            clean_entity_field = clean_text(entity_field)
            if clean_entity_field and isinstance(columns, list):
                aliases.setdefault(clean_entity_field, [])
                aliases[clean_entity_field].extend(clean_text(column) for column in columns)

    if aliases:
        return aliases

    fields = mapping_config.get("fields")
    if isinstance(fields, list):
        legacy_aliases: dict[str, list[str]] = {}
        for field in fields:
            if not isinstance(field, dict):
                continue
            target_field = clean_text(field.get("entity_field") or field.get("target_field"))
            field_source_columns = field.get("source_columns") or []
            if target_field:
                legacy_aliases[target_field] = [clean_text(column) for column in field_source_columns]
        return legacy_aliases

    return {
        field: [clean_text(column) for column in columns]
        for field, columns in mapping_config.items()
        if isinstance(columns, list)
    }


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text = clean_text(value)
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return datetime.fromisoformat(text)
